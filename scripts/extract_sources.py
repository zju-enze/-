#!/usr/bin/env python3
"""Extract bounded evidence from local thesis source materials."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any


TEXT_EXTENSIONS = {
    ".tex",
    ".bib",
    ".md",
    ".markdown",
    ".rst",
    ".txt",
    ".py",
    ".r",
    ".jl",
    ".m",
    ".c",
    ".cc",
    ".cpp",
    ".h",
    ".hpp",
    ".java",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".go",
    ".rs",
    ".sh",
    ".ps1",
    ".sql",
    ".yaml",
    ".yml",
    ".json",
    ".toml",
    ".ini",
    ".cfg",
}


def trim(text: str, limit: int) -> tuple[str, bool]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    if len(normalized) <= limit:
        return normalized, False
    return normalized[:limit], True


def read_text(path: Path, limit: int) -> tuple[str, dict[str, Any]]:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            excerpt, truncated = trim(text, limit)
            return excerpt, {"encoding": encoding, "truncated": truncated}
        except UnicodeDecodeError:
            continue
    return "", {"error": "unable to decode text"}


def extract_pdf(path: Path, limit: int) -> tuple[str, dict[str, Any]]:
    errors: list[str] = []
    try:
        import fitz  # type: ignore

        doc = fitz.open(path)
        parts = []
        for page in doc:
            parts.append(page.get_text())
            if sum(len(part) for part in parts) >= limit:
                break
        text, truncated = trim("\n".join(parts), limit)
        return text, {"pages": doc.page_count, "engine": "fitz", "truncated": truncated}
    except Exception as exc:  # noqa: BLE001
        errors.append(f"fitz: {exc}")
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
            if sum(len(part) for part in parts) >= limit:
                break
        text, truncated = trim("\n".join(parts), limit)
        return text, {"pages": len(reader.pages), "engine": "pypdf", "truncated": truncated}
    except Exception as exc:  # noqa: BLE001
        errors.append(f"pypdf: {exc}")
    return "", {"error": "; ".join(errors) or "no PDF extractor available"}


def extract_docx(path: Path, limit: int) -> tuple[str, dict[str, Any]]:
    try:
        import docx  # type: ignore

        document = docx.Document(str(path))
        paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
        text, truncated = trim("\n".join(paragraphs), limit)
        return text, {"paragraphs": len(paragraphs), "engine": "python-docx", "truncated": truncated}
    except Exception as exc:  # noqa: BLE001
        return "", {"error": str(exc)}


def extract_spreadsheet(path: Path, limit: int) -> tuple[str, dict[str, Any]]:
    if path.suffix.lower() in {".csv", ".tsv"}:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        lines = []
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh, delimiter=delimiter)
                for row_number, row in enumerate(reader, start=1):
                    lines.append(" | ".join(row))
                    if row_number >= 50 or sum(len(line) for line in lines) >= limit:
                        break
            text, truncated = trim("\n".join(lines), limit)
            return text, {"rows_sampled": len(lines), "truncated": truncated}
        except Exception as exc:  # noqa: BLE001
            return "", {"error": str(exc)}
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"[Sheet] {sheet.title}")
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    lines.append(" | ".join(values))
                if row_number >= 30 or sum(len(line) for line in lines) >= limit:
                    break
            if sum(len(line) for line in lines) >= limit:
                break
        text, truncated = trim("\n".join(lines), limit)
        return text, {"sheets": workbook.sheetnames, "engine": "openpyxl", "truncated": truncated}
    except Exception as exc:  # noqa: BLE001
        return "", {"error": str(exc)}


def extract_image(path: Path) -> tuple[str, dict[str, Any]]:
    try:
        from PIL import Image  # type: ignore

        with Image.open(path) as image:
            return "", {"format": image.format, "width": image.width, "height": image.height, "mode": image.mode}
    except Exception as exc:  # noqa: BLE001
        return "", {"error": str(exc)}


def extract_file(root: Path, item: dict[str, Any], limit: int) -> dict[str, Any]:
    path = root / item["path"]
    suffix = path.suffix.lower()
    category = item.get("category", "other")
    if suffix == ".pdf":
        text, meta = extract_pdf(path, limit)
    elif suffix == ".docx":
        text, meta = extract_docx(path, limit)
    elif suffix in {".xlsx", ".xls", ".csv", ".tsv"}:
        text, meta = extract_spreadsheet(path, limit)
    elif category == "image":
        text, meta = extract_image(path)
    elif suffix in TEXT_EXTENSIONS or category in {"latex", "markdown", "bibliography", "code", "config", "data"}:
        text, meta = read_text(path, limit)
    else:
        text, meta = "", {"note": "unsupported or binary file; inventory only"}
    return {
        "path": item["path"],
        "category": category,
        "size_bytes": item.get("size_bytes"),
        "metadata": meta,
        "excerpt": text,
    }


def markdown_report(data: dict[str, Any], extracted: list[dict[str, Any]]) -> str:
    lines = [
        "# Thesis Source Evidence",
        "",
        "This file is a source-grounding aid. Use it to write the thesis; do not treat unsupported files as evidence beyond their metadata.",
        "",
        "## Inventory Summary",
        "",
    ]
    for category, count in data.get("counts", {}).items():
        lines.append(f"- {category}: {count}")
    git = data.get("git")
    if git:
        lines.extend(["", "## Git Summary", ""])
        for key in ("top_level", "branch", "head", "status_short"):
            value = git.get(key)
            if value:
                lines.append(f"- {key}: `{value}`")
        commits = git.get("recent_commits") or []
        if commits:
            lines.append("- recent commits:")
            for commit in commits[:20]:
                lines.append(f"  - {commit}")
    lines.extend(["", "## Extracted Files", ""])
    for item in extracted:
        lines.append(f"### {item['path']}")
        lines.append("")
        lines.append(f"- category: {item['category']}")
        lines.append(f"- size_bytes: {item.get('size_bytes')}")
        metadata = item.get("metadata") or {}
        if metadata:
            lines.append(f"- metadata: `{json.dumps(metadata, ensure_ascii=False)}`")
        excerpt = item.get("excerpt") or ""
        if excerpt:
            lines.extend(["", "```text", excerpt.strip(), "```", ""])
        else:
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("root", nargs="?", default=".", help="Directory to extract from.")
    parser.add_argument("--inventory", default=None, help="Inventory JSON path.")
    parser.add_argument("--output", default="thesis_sources.md", help="Markdown report path.")
    parser.add_argument("--json-output", default=None, help="Optional JSON extraction output.")
    parser.add_argument("--limit", type=int, default=8000, help="Maximum extracted characters per file.")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    if args.inventory:
        data = json.loads(Path(args.inventory).read_text(encoding="utf-8"))
    else:
        from inventory_sources import inventory

        data = inventory(root)
    extracted = [extract_file(root, item, args.limit) for item in data.get("files", [])]
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(markdown_report(data, extracted), encoding="utf-8")
    if args.json_output:
        json_output = Path(args.json_output)
        json_output.parent.mkdir(parents=True, exist_ok=True)
        json_output.write_text(json.dumps({"inventory": data, "extracted": extracted}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {output} with {len(extracted)} extracted entries.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
